import os
import json
import openpyxl
import psycopg2
from datetime import date
from dotenv import load_dotenv,find_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font

load_dotenv(find_dotenv())
f = open('config.json')
jdata = json.load(f)

def getMapping(status_codes):
  status_codes_mapping = { "1":"G_INITIATED", "101":"PG_INITIATED","125":"PG_FAILED","150":"PG_PENDING","199":"PG_SUCCESS","201":"BBPS_INITIATED","225":"BBPS_FAILED","250":"BBPS_PENDING","299":"BBPS_SUCCESS","301":"SETTLE_INITIATED","325":"SETTLE_FAILED","350":"SETTLE_PENDING","399":"SETTLE_SUCCESS"}
  return status_codes_mapping.get(status_codes)

def export_to_excel(connection, query_string, headings, filepath,status_code_colnum,listofprecision):
    cursor = connection.cursor()
    cursor.execute(query_string)
    data = cursor.fetchall() 
    data1=[]
    for tup in data:
      lst = list(tup)
      lst[status_code_colnum] = getMapping(str(tup[status_code_colnum]))
      for l in listofprecision:
        if lst[l] is not None:
          lst[l] = round(lst[l],2)  
      data1.append(lst)
      
    cursor.close()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.row_dimensions[1].font = Font(bold = True)
    #header
    for colno, heading in enumerate(headings,start = 1):
        sheet.cell(row = 1, column = colno).value = heading
    #rows
    for rowno, row in enumerate(data1, start = 2):
        for colno, cell_value in enumerate(row, start = 1):
            sheet.cell(row = rowno, column = colno).value = cell_value
    wb.save(filepath)
    print("File Created in your workspace")

def getDatabaseConnection():
  try:
    conn = psycopg2.connect(
      dbname=jdata['conn']['PAYMENT_DB_NAME'], 
      user=jdata['conn']['PAYMENT_DB_USER'], 
      password=jdata['conn']['PAYMENT_DB_PASSWORD'], 
      host=jdata['conn']['DB_HOST'], 
      port=jdata['conn']['DB_PORT'], 
      )
    print("Connected to Database")
    return conn
  except Exception as e:
    print("Error is database connection ",e)

def main():
    COLS = ["company_id","loan_id","amount","pg_charges","payment_mode","pg_success_datetime","bbps_success_datetime","status_code_id","biller_id","biller_category","payment_message","settle_date","created","updated","charges_by_client","cg_charges","allocation_month","cg_charges_gst","user_id"]
    STATUS_CODE_COL_NUM = 7;
    FILENAME = f"payments_transactions_{date.today()}.xlsx"
    LIST_PRECISION_SETTLEMENT = [2,3,14,15,17]
    QUERY_WHERE_CONDITION_STATUS_CODE = 299
    try:
      conn = getDatabaseConnection()
      joined_string = ",".join(COLS) 
      query_string = "select " + joined_string + " from transaction where {}=299".format(QUERY_WHERE_CONDITION_STATUS_CODE);
      export_to_excel(conn, query_string, COLS,FILENAME,STATUS_CODE_COL_NUM,LIST_PRECISION_SETTLEMENT)
                               
    except Exception as e:
      print(e)

main()